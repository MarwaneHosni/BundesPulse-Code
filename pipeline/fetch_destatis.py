"""Fetch official Destatis population data and stage it for the snapshot build.

Downloads two official Destatis statistical reports (GENESIS-Online table 12411,
`Bevoelkerungsfortschreibung`, Zensus-2022 basis) and extracts the embedded
CSV tables:

* per-Bundesland (and Germany) population for the report years 2023 and 2024,
  with the officially published growth rate and age-group counts;
* Germany time series 1990-2024.

The parsed data is written as staging CSVs under ``data/processed/destatis/``
which ``pipeline/build_data.py`` then loads into DuckDB. Downloaded files stay
under ``data/raw/destatis/`` (both directories are gitignored).

Nothing here runs at request time - this is the offline data-build step.
"""

from __future__ import annotations

import csv
import os
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = Path(
    os.environ.get("BUNDESPULSE_RAW_DIR", REPO_ROOT / "data" / "raw" / "destatis")
)
OUT_DIR = Path(
    os.environ.get(
        "BUNDESPULSE_PROCESSED_DIR", REPO_ROOT / "data" / "processed" / "destatis"
    )
)

# "Statistischer Bericht Bevoelkerungsfortschreibung (Zensus 2022)" - table 12411.
REPORTS = {
    "bev_2023": {
        "url": "https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/"
        "Bevoelkerungsstand/Publikationen/Downloads-Bevoelkerungsstand/"
        "statistischer-bericht-bevoelkerungsfortschreibung-zensus-2022-jaehrlich-"
        "5124108237005.xlsx?__blob=publicationFile&v=4",
        "file": "12411_bevoelkerung_2023.xlsx",
    },
    "bev_2024": {
        "url": "https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/"
        "Bevoelkerungsstand/Publikationen/Downloads-Bevoelkerungsstand/"
        "statistischer-bericht-bevoelkerungsfortschreibung-zensus-2022-jaehrlich-"
        "5124108247005.xlsx?__blob=publicationFile&v=4",
        "file": "12411_bevoelkerung_2024.xlsx",
    },
}

# Official federal-state (Bundesland) AGS codes used as canonical region ids.
LAND_AGS = {
    "Baden-Württemberg": "01",
    "Bayern": "02",
    "Berlin": "03",
    "Brandenburg": "04",
    "Bremen": "05",
    "Hamburg": "06",
    "Hessen": "07",
    "Mecklenburg-Vorpommern": "08",
    "Niedersachsen": "09",
    "Nordrhein-Westfalen": "10",
    "Rheinland-Pfalz": "11",
    "Saarland": "12",
    "Sachsen": "13",
    "Sachsen-Anhalt": "14",
    "Schleswig-Holstein": "15",
    "Thüringen": "16",
}
COUNTRY_ID = "DE"


class FetchError(RuntimeError):
    """A downloaded report is unusable; fail the fetch."""


def download(url: str, dest: Path) -> None:
    if dest.exists():
        print(f"  using cached {dest.name}")
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp, open(dest, "wb") as out:
        out.write(resp.read())
    print(f"  downloaded {dest.name} ({dest.stat().st_size:,} bytes)")


def _row_dict(row: tuple) -> dict:
    """Map a bare-tuple row to a dict using column-position constants."""
    return {
        # csv-12411-08: code, label, year, Bundesland, pop, growth, ... u18, ... 65+
        "year": _cell(row, 2),
        "region": _region_name(row, 3),
        "pop": _cell(row, 4),
        "growth": _cell(row, 5),
        "u18": _cell(row, 13),
        "g65": _cell(row, 16),
    }


def _cell(row: tuple, idx: int):
    if idx < len(row) and row[idx] is not None:
        return row[idx]
    return None


def _region_name(row: tuple, idx: int) -> str:
    raw = _cell(row, idx) or ""
    name = str(raw).strip()
    # Report uses underscore forms for aggregates such as "Frueheres_Bundesgebiet".
    return {"Insgesamt": COUNTRY_ID, "Deutschland": COUNTRY_ID}.get(name, name)


def _parse_laender_rows(ws, berichtsjahr: int) -> list[dict]:
    """Parse the per-Bundesland sheet csv-12411-08 for one report year."""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        rec = _row_dict(row)
        if rec["region"] == COUNTRY_ID or rec["region"] in LAND_AGS:
            years = str(rec["year"] or "").strip()
            if str(berichtsjahr) in (years,):
                rows.append(rec)
    return rows


def _parse_deutschland_series(ws) -> list[dict]:
    """Parse the Germany time series sheet csv-12411-03 (year -> pop, growth)."""
    series = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        year = _cell(row, 2)
        pop = _to_float(_cell(row, 3))
        growth = _to_float(_cell(row, 4))
        if year is None:
            continue
        series.append({"year": int(year), "pop": pop, "growth": growth})
    return series


def _to_float(v):
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _region_to_id(name: str) -> str:
    """Map a report region name to its canonical region_id."""
    if name == COUNTRY_ID:
        return COUNTRY_ID
    return LAND_AGS.get(name, name)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    reports = {}  # key -> {"year","rows": {region_id: rec}, "series": [...]}
    for key, spec in REPORTS.items():
        raw_file = RAW_DIR / spec["file"]
        print(f"[{key}] download report ...")
        download(spec["url"], raw_file)
        wb = openpyxl.load_workbook(raw_file, read_only=True, data_only=True)
        if "csv-12411-08" not in wb.sheetnames or "csv-12411-03" not in wb.sheetnames:
            raise FetchError(f"report {key}: unexpected sheet layout: {wb.sheetnames}")
        laender = _parse_laender_rows(
            wb["csv-12411-08"], berichtsjahr=2023 if key == "bev_2023" else 2024
        )
        series = _parse_deutschland_series(wb["csv-12411-03"])
        wb.close()
        if not laender:
            raise FetchError(f"report {key}: no Bundesland rows found")
        by_region = {rec["region"]: rec for rec in laender}
        reports[key] = {
            "year": 2023 if key == "bev_2023" else 2024,
            "rows": by_region,
            "series": series,
        }
        print(
            f"  parsed {len(by_region)} regions, year {reports[key]['year']}, {len(series)} DE series rows"
        )

    if len(reports["bev_2023"]["rows"]) != 17:
        print(
            f"  WARNING: 2023 report has {len(reports['bev_2023']['rows'])} regions (expected 17)"
        )

    # ---- regions catalogue ----
    regions = [("DE", "Deutschland", "bund", None, None)]
    regions += [
        (code, name, "bundesland", "DE", None) for name, code in LAND_AGS.items()
    ]

    # ---- indicators ----
    indicators = [
        (
            1,
            "pop_total",
            "Bevölkerungsstand",
            "Demography",
            "persons",
            "Bevölkerungsstand am 31.12. (Bevölkerungsfortschreibung, Basis Zensus 2022)",
            "raw",
        ),
        (
            2,
            "pop_growth",
            "Bevölkerungswachstum",
            "Demography",
            "percent",
            "Vom Statistischen Bundesamt berechnete Wachstumsrate gegenüber dem Vorjahr",
            "raw",
        ),
        (
            3,
            "pop_share_65plus",
            "Anteil 65-Jährige und Ältere",
            "Demography",
            "percent",
            "Anteil der 65-Jährigen und Älteren an der Bevölkerung (abgeleitet)",
            "derived",
        ),
        (
            4,
            "pop_share_under18",
            "Anteil unter 18-Jährige",
            "Demography",
            "percent",
            "Anteil der unter 18-Jährigen an der Bevölkerung (abgeleitet)",
            "derived",
        ),
    ]

    # ---- sources (one entry per downloaded report) ----
    today = datetime.now(timezone.utc).date().isoformat()
    sources = [
        (
            1,
            "Statistisches Bundesamt (Destatis)",
            "Bevölkerungsfortschreibung 2023 (Statistischer Bericht 12411)",
            REPORTS["bev_2023"]["url"],
            today,
        ),
        (
            2,
            "Statistisches Bundesamt (Destatis)",
            "Bevölkerungsfortschreibung 2024 (Statistischer Bericht 12411)",
            REPORTS["bev_2024"]["url"],
            today,
        ),
    ]

    # ---- observations ----
    obs = []  # (region_id, indicator_id, period, value, source_id)

    def add(region: str, indicator_id: int, period: int, value, source_id: int) -> None:
        if value is None:
            return
        obs.append((region, indicator_id, period, value, source_id))

    for key, rep in reports.items():
        year, rows, source_id = rep["year"], rep["rows"], 1 if key == "bev_2023" else 2
        for region_name, rec in rows.items():
            region = _region_to_id(region_name)
            pop = _to_float(rec["pop"])
            growth = _to_float(rec["growth"])
            u18 = _to_float(rec["u18"])
            g65 = _to_float(rec["g65"])
            add(region, 1, year, pop, source_id)
            add(region, 2, year, growth, source_id)
            if pop:
                add(
                    region,
                    3,
                    year,
                    round(g65 / pop * 100.0, 2) if g65 is not None else None,
                    source_id,
                )
                add(
                    region,
                    4,
                    year,
                    round(u18 / pop * 100.0, 2) if u18 is not None else None,
                    source_id,
                )

    # Germany long time series (1990-2024) from the newest report.
    for point in reports["bev_2024"]["series"]:
        add(COUNTRY_ID, 1, point["year"], point["pop"], 2)
        add(COUNTRY_ID, 2, point["year"], point["growth"], 2)

    # ---- write staging CSVs (schema matches build_data.py) ----
    _write_csv(
        OUT_DIR / "regions.csv",
        ["region_id", "name", "type", "parent_id", "area"],
        regions,
    )
    _write_csv(
        OUT_DIR / "indicators.csv",
        [
            "indicator_id",
            "slug",
            "name",
            "category",
            "unit",
            "description",
            "raw_or_derived",
        ],
        indicators,
    )
    _write_csv(
        OUT_DIR / "observations.csv",
        ["region_id", "indicator_id", "period", "value", "source_id"],
        obs,
    )
    _write_csv(
        OUT_DIR / "sources.csv",
        ["source_id", "provider", "dataset", "url", "retrieval_date"],
        sources,
    )

    print(f"\nwrote staging CSVs to {OUT_DIR}")
    print(
        f"  regions: {len(regions)} | indicators: {len(indicators)} | "
        f"observations: {len(obs)} | sources: {len(sources)}"
    )


def _write_csv(path: Path, header: list[str], rows: list[tuple]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        writer.writerows(rows)


if __name__ == "__main__":
    main()
