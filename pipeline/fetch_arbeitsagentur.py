"""Fetch labour-market data from the Bundesagentur fuer Arbeit and stage it.

Sources (official BA Statistik, "Einzelhefte" downloads, no credentials):

* *Arbeitslose* (persons) and *Arbeitslosenquote* (%) per Kreis,
  December 2025 (`arbeitslose-quoten-dlk`).
* *Sozialversicherungspflichtig Beschaeftigte am Arbeitsort* (persons) per
  Bundesland, June 2022 (`gemband-dlk`).

Staging CSVs are written to ``data/processed/arbeitsagentur/`` (gitignored),
in the same schema used by ``pipeline/build_data.py``.
"""

from __future__ import annotations

import csv
import io
import os
import re
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone
from datetime import datetime as _dt
from pathlib import Path

import openpyxl
from pyxlsb import open_workbook

from pipeline.source_regions import LAND_AGS, kreis_parent

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = REPO_ROOT / "data" / "raw" / "sources"
OUT_DIR = Path(
    os.environ.get(
        "BUNDESPULSE_PROCESSED_DIR", REPO_ROOT / "data" / "processed" / "arbeitsagentur"
    )
)

# BA Statistik "Einzelheftsuche" downloads.
ARBEITSLOSE_URL = (
    "https://statistik.arbeitsagentur.de/Statistikdaten/Detail/202512/iiia4/"
    "gemeinde-arbeitslose-quoten/arbeitslose-quoten-dlk-0-202512-zip.zip?__blob=publicationFile&v=1"
)
ARBEITSLOSE_FILE = "ba_arbeitslose_quoten.zip"

SVB_URL = (
    "https://statistik.arbeitsagentur.de/Statistikdaten/Detail/202206/iiia6/"
    "beschaeftigung-sozbe-gemband/gemband-dlk-0-202206-zip.zip?__blob=publicationFile&v=2"
)
SVB_FILE = "ba_svb.zip"

HEADERS = {"User-Agent": "Mozilla/5.0"}

# global indicator ids: 1-4 = population (destatis), 5+ = labour market.
ID_UNEMP, ID_UNEMP_RATE, ID_EMP = 5, 6, 7


def download(url: str, dest: Path) -> Path:
    if not dest.exists():
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=120) as resp, open(dest, "wb") as out:
            out.write(resp.read())
        print(f"  downloaded {dest.name} ({dest.stat().st_size:,} bytes)")
    else:
        print(f"  using cached {dest.name}")
    return dest


def _as_float(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return float(s)
    return None


def _parse_arbeitslosen_heft(zip_path: Path, measure: str) -> dict[str, float]:
    """Parse the 'Übersicht_Kreise' sheet of the Arbeitslose/Arbeitslosenquoten heft.

    Returns {ags: Dezember-2025 value}.
    """
    zf = zipfile.ZipFile(zip_path)
    member = (
        "202512_Heft_pol_SGBI_Arbeitslose.xlsx"
        if measure == "unemp"
        else "202512_Heft_pol_SGBI_ArbeitslosenQuoten.xlsx"
    )
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read(member)), read_only=True, data_only=True
    )
    ws = wb["Übersicht_Kreise"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # The last monthly column is Dezember 2025. Locate it in the header row.
    header_idx = last_month_idx = None
    for i, r in enumerate(rows):
        for j, c in enumerate(r):
            is_dec_2025 = (isinstance(c, _dt) and c.year == 2025 and c.month == 12) or (
                isinstance(c, str) and c.strip().startswith("2025-12")
            )
            if is_dec_2025:
                header_idx, last_month_idx = i, j
                break
        if header_idx is not None:
            break
    if header_idx is None:
        raise RuntimeError("Arbeitslosenheft: Dezember-2025 header column not found")

    values: dict[str, float] = {}
    for r in rows[header_idx + 1 :]:
        if not r or not r[0]:
            continue
        m = re.fullmatch(r"\s*(\d{5})\s+(.+)", str(r[0]).strip())
        if not m:
            continue
        cell = r[last_month_idx] if last_month_idx < len(r) else None
        fv = _as_float(cell)
        if fv is not None:
            values[m.group(1)] = fv
    return values


def _parse_svb_laender(zip_path: Path) -> dict[str, float]:
    """Parse the 'Länder' sheet of the svB Gemeindedaten file (Arbeitsort)."""
    zf = zipfile.ZipFile(zip_path)
    raw = zf.read("gemband_dlk_0.xlsb")
    values: dict[str, float] = {}
    with open_workbook(io.BytesIO(raw)) as wb:
        ws = wb.get_sheet("Länder")
        for row in ws.rows():
            cells = [c.v for c in row]
            if not cells or cells[0] is None or not str(cells[0]).strip().isdigit():
                continue
            if len(cells) < 11:
                continue
            arbeitsort = _as_float(cells[10])
            if arbeitsort is None:
                continue
            land_name = str(cells[1]).strip()
            land_id = LAND_AGS.get(land_name)
            if land_id:
                values[land_id] = arbeitsort
    return values


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now(timezone.utc).date().isoformat()

    print("- BA: Arbeitslose / Arbeitslosenquote (Kreise, Dez. 2025)")
    al_zip = download(ARBEITSLOSE_URL, RAW_DIR / ARBEITSLOSE_FILE)
    arbeitslose = _parse_arbeitslosen_heft(al_zip, "unemp")
    quoten = _parse_arbeitslosen_heft(al_zip, "unemp_rate")
    kreis_names = _kreis_names(al_zip)
    print(f"  Kreise: arbeitslose={len(arbeitslose)} quoten={len(quoten)}")

    print("- BA: svB Beschaeftigte (Laender, Jun. 2022)")
    svb_zip = download(SVB_URL, RAW_DIR / SVB_FILE)
    emp = _parse_svb_laender(svb_zip)
    print(f"  Laender: emp_social={len(emp)}")

    # indicators (global ids: 1-4 = population, 5-6 BA arbeitslose, 7 employment)
    indicators = [
        (
            ID_UNEMP,
            "unemp",
            "Arbeitslose",
            "Employment",
            "persons",
            "Arbeitslose insgesamt (BA-Statistik), Stand Dezember 2025",
            "raw",
        ),
        (
            ID_UNEMP_RATE,
            "unemp_rate",
            "Arbeitslosenquote",
            "Employment",
            "percent",
            "Arbeitslosenquote in Prozent (aller zivilen Erwerbspersonen), Dezember 2025",
            "raw",
        ),
        (
            ID_EMP,
            "emp_social",
            "Sozialversicherungspflichtig Beschäftigte am Arbeitsort",
            "Employment",
            "persons",
            "svB am Arbeitsort (BA Beschäftigungsstatistik), Juni 2022",
            "raw",
        ),
    ]

    sources = [
        (
            1,
            "Bundesagentur für Arbeit",
            "Arbeitslose und Arbeitslosenquoten nach Kreisen (Dezember 2025)",
            ARBEITSLOSE_URL,
            today,
        ),
        (
            2,
            "Bundesagentur für Arbeit",
            "svB Beschäftigte am Arbeitsort nach Ländern (Juni 2022)",
            SVB_URL,
            today,
        ),
    ]

    obs = []
    for ags, val in arbeitslose.items():
        if val is not None:
            obs.append((ags, ID_UNEMP, 2025, val, 1))
    for ags, val in quoten.items():
        if val is not None:
            obs.append((ags, ID_UNEMP_RATE, 2025, val, 1))
    for land_id, val in emp.items():
        if val is not None:
            obs.append((land_id, ID_EMP, 2022, val, 2))

    regions = sorted(
        (ags, kreis_names[ags], "kreis", kreis_parent(ags), None) for ags in kreis_names
    )

    _write_csv(
        OUT_DIR / "regions.csv",
        ["region_id", "name", "type", "parent_id", "area"],
        regions,
    )
    _write_csv(
        OUT_DIR / "indicators.csv",
        [
            "indicator_id",
            "slug",
            "name",
            "category",
            "unit",
            "description",
            "raw_or_derived",
        ],
        indicators,
    )
    _write_csv(
        OUT_DIR / "observations.csv",
        ["region_id", "indicator_id", "period", "value", "source_id"],
        obs,
    )
    _write_csv(
        OUT_DIR / "sources.csv",
        ["source_id", "provider", "dataset", "url", "retrieval_date"],
        sources,
    )

    print(f"\nwrote staging CSVs to {OUT_DIR}")
    print(
        f"  regions(kreise): {len(regions)} | indicators: {len(indicators)} | observations: {len(obs)}"
    )


def _kreis_names(zip_path: Path) -> dict[str, str]:
    zf = zipfile.ZipFile(zip_path)
    wb = openpyxl.load_workbook(
        io.BytesIO(zf.read("202512_Heft_pol_SGBI_Arbeitslose.xlsx")),
        read_only=True,
        data_only=True,
    )
    ws = wb["Übersicht_Kreise"]
    names = {}
    for r in ws.iter_rows(values_only=True):
        if r and r[0]:
            m = re.fullmatch(r"\s*(\d{5})\s+(.+)", str(r[0]).strip())
            if m:
                names[m.group(1)] = m.group(2).strip()
    wb.close()
    return names


def _write_csv(path: Path, header: list[str], rows: list[tuple]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


if __name__ == "__main__":
    main()
